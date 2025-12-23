from __future__ import annotations

import io
import os
import re
import shutil
import tempfile
import traceback
from datetime import time
from typing import Optional, Tuple, List

import streamlit as st

# ============================================================
# ✅ Page config (must be first Streamlit command)
# ============================================================
st.set_page_config(page_title="EDD MIS Chatbot", layout="wide")


# ============================================================
# ✅ Lazy dependency loader (prevents "Oh no" on missing libs)
# ============================================================
pd = None
np = None
PatternFill = Font = Alignment = Border = Side = ColorScaleRule = None


def _ensure_deps() -> bool:
    """Import heavy deps safely. If fails, show error in UI (no crash)."""
    global pd, np, PatternFill, Font, Alignment, Border, Side, ColorScaleRule
    if pd is not None and np is not None:
        return True
    try:
        import pandas as _pd
        import numpy as _np
        from openpyxl.styles import PatternFill as _PatternFill, Font as _Font, Alignment as _Alignment, Border as _Border, Side as _Side
        from openpyxl.formatting.rule import ColorScaleRule as _ColorScaleRule

        pd = _pd
        np = _np
        PatternFill, Font, Alignment, Border, Side, ColorScaleRule = (
            _PatternFill, _Font, _Alignment, _Border, _Side, _ColorScaleRule
        )
        return True
    except Exception:
        st.error("❌ Dependency import failed. Fix requirements.txt.")
        st.code(traceback.format_exc())
        return False


# ============================================================
# ✅ FAST SUMMARY BUILDERS (OPTIMIZED)
# ============================================================
def _format_percent_count(pct: float, cnt: int) -> str:
    try:
        cnt_i = int(cnt)
    except Exception:
        cnt_i = 0
    if cnt_i <= 0:
        return "0% (0)"
    try:
        pct_f = round(float(pct), 2)
    except Exception:
        pct_f = 0.0
    return f"{pct_f}% ({cnt_i})"


def _build_summary_fast(df):
    """
    Builds the same summary as your original code but MUCH faster.
    Avoids nested df[(...)] filters; uses groupby/unstack cubes.
    """
    if not _ensure_deps():
        st.stop()

    # Keep same week ordering as original:
    weekly_total_df = df.groupby("Week_Label").size().reset_index(name="Picked Volume")
    weeks = weekly_total_df["Week_Label"].tolist()

    weekly_total = (
        weekly_total_df.set_index("Week_Label")["Picked Volume"]
        .reindex(weeks)
        .fillna(0)
        .astype(int)
    )

    zones = sorted(df["BKG_Zone"].dropna().unique())
    all_modes = sorted(df["TPTR_Mode"].dropna().unique())
    all_statuses = sorted(df["CN_Current_Status"].dropna().unique())
    all_business_types = sorted(df["BUSINESS_TYPE"].dropna().unique())

    # zone x week
    zone_counts = (
        df.groupby(["BKG_Zone", "Week_Label"]).size().unstack(fill_value=0)
        .reindex(index=zones, columns=weeks, fill_value=0)
        .astype(int)
    )

    # zone x business_type x week
    bt_counts = df.groupby(["BKG_Zone", "BUSINESS_TYPE", "Week_Label"]).size().unstack(fill_value=0)
    bt_counts = bt_counts.reindex(
        pd.MultiIndex.from_product([zones, all_business_types], names=["BKG_Zone", "BUSINESS_TYPE"]),
        fill_value=0,
    ).reindex(columns=weeks, fill_value=0).astype(int)

    # zone x mode x week
    mode_counts = df.groupby(["BKG_Zone", "TPTR_Mode", "Week_Label"]).size().unstack(fill_value=0)
    mode_counts = mode_counts.reindex(
        pd.MultiIndex.from_product([zones, all_modes], names=["BKG_Zone", "TPTR_Mode"]),
        fill_value=0,
    ).reindex(columns=weeks, fill_value=0).astype(int)

    # OTA cube
    ota_counts = (
        df[df["ON_TIME_ARRIVAL"] == "Yes"]
        .groupby(["BKG_Zone", "TPTR_Mode", "Week_Label"]).size().unstack(fill_value=0)
    )
    ota_counts = ota_counts.reindex(
        pd.MultiIndex.from_product([zones, all_modes], names=["BKG_Zone", "TPTR_Mode"]),
        fill_value=0,
    ).reindex(columns=weeks, fill_value=0).astype(int)

    # OTD cube
    otd_counts = (
        df[df["ON_TIME_DELIVERY"] == "Yes"]
        .groupby(["BKG_Zone", "TPTR_Mode", "Week_Label"]).size().unstack(fill_value=0)
    )
    otd_counts = otd_counts.reindex(
        pd.MultiIndex.from_product([zones, all_modes], names=["BKG_Zone", "TPTR_Mode"]),
        fill_value=0,
    ).reindex(columns=weeks, fill_value=0).astype(int)

    # NDR not available cube
    ndr_blank = df["NDR_Remark"].isna() | (df["NDR_Remark"].astype(str).str.strip() == "")
    ndr_mask = (df["CN_Current_Status"] == "Ware house Destination") & ndr_blank
    ndr_counts = (
        df[ndr_mask].groupby(["BKG_Zone", "Week_Label"]).size().unstack(fill_value=0)
        .reindex(index=zones, columns=weeks, fill_value=0)
        .astype(int)
    )

    # Status cube
    status_counts = df.groupby(["BKG_Zone", "CN_Current_Status", "Week_Label"]).size().unstack(fill_value=0)
    status_counts = status_counts.reindex(
        pd.MultiIndex.from_product([zones, all_statuses], names=["BKG_Zone", "CN_Current_Status"]),
        fill_value=0,
    ).reindex(columns=weeks, fill_value=0).astype(int)

    idx: List[str] = []
    rows: List[List[object]] = []

    # Picked Volume
    idx.append("Picked Volume")
    rows.append(weekly_total.values.tolist())

    wt = weekly_total.values.astype(float)

    for zone in zones:
        zc = zone_counts.loc[zone].values.astype(int)

        # Picked Vol Zone %
        pct = np.where(wt != 0, (zc / wt) * 100.0, 0.0)
        idx.append(f"Picked Vol. Zone {zone} %")
        rows.append([_format_percent_count(p, c) for p, c in zip(pct, zc)])

        # BUSINESS TYPE BREAKDOWN header
        idx.append(f"BUSINESS TYPE BREAKDOWN__{zone}")
        rows.append(["" for _ in weeks])

        for bt in all_business_types:
            bc = bt_counts.loc[(zone, bt)].values.astype(int)
            denom = zc.astype(float)
            pct = np.where(denom != 0, (bc / denom) * 100.0, 0.0)
            idx.append(f"   {bt}__{zone}")
            rows.append([_format_percent_count(p, c) for p, c in zip(pct, bc)])

        # TPTR Mode blocks
        for mode in all_modes:
            mc = mode_counts.loc[(zone, mode)].values.astype(int)

            denom = zc.astype(float)
            pct = np.where(denom != 0, (mc / denom) * 100.0, 0.0)
            idx.append(f"TPTR Mode {mode}__{zone}")
            rows.append([_format_percent_count(p, c) for p, c in zip(pct, mc)])

            oc = ota_counts.loc[(zone, mode)].values.astype(int)
            denom = mc.astype(float)
            pct = np.where(denom != 0, (oc / denom) * 100.0, 0.0)
            idx.append(f"{mode} On Time Arrival__{zone}")
            rows.append([_format_percent_count(p, c) for p, c in zip(pct, oc)])

            dc = otd_counts.loc[(zone, mode)].values.astype(int)
            denom = mc.astype(float)
            pct = np.where(denom != 0, (dc / denom) * 100.0, 0.0)
            idx.append(f"{mode} On Time Delivery__{zone}")
            rows.append([_format_percent_count(p, c) for p, c in zip(pct, dc)])

        # NDR not available
        nc = ndr_counts.loc[zone].values.astype(int)
        denom = zc.astype(float)
        pct = np.where(denom != 0, (nc / denom) * 100.0, 0.0)
        idx.append(f"NDR not available__{zone}")
        rows.append([_format_percent_count(p, c) for p, c in zip(pct, nc)])

        # CN Status Breakdown header
        idx.append(f"CN Status Breakdown__{zone}")
        rows.append(["" for _ in weeks])

        for status in all_statuses:
            sc = status_counts.loc[(zone, status)].values.astype(int)
            denom = zc.astype(float)
            pct = np.where(denom != 0, (sc / denom) * 100.0, 0.0)
            idx.append(f"   {status}__{zone}")
            rows.append([_format_percent_count(p, c) for p, c in zip(pct, sc)])

    summary = pd.DataFrame(rows, index=idx, columns=weeks)
    # Clean labels like original
    summary.index = summary.index.to_series().astype(str).str.replace(r"__(.*)$", "", regex=True).values
    return summary


def _apply_row_grouping(ws) -> None:
    """Your original grouping logic, optimized (cache column A reads)."""
    max_row = ws.max_row
    colA = [None] + [ws.cell(row=r, column=1).value for r in range(1, max_row + 1)]

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    row = 2
    while row <= max_row:
        val = colA[row]
        if isinstance(val, str) and val.startswith("Picked Vol. Zone"):
            ws.row_dimensions[row].outline_level = 1
            ws.row_dimensions[row].collapsed = True

            zone_start = row + 1
            r = zone_start
            while r <= max_row and not (
                isinstance(colA[r], str) and str(colA[r]).startswith("Picked Vol. Zone")
            ):
                r += 1
            zone_end = r - 1

            if zone_end >= zone_start:
                ws.row_dimensions.group(zone_start, zone_end, hidden=True, outline_level=2)

            sub = zone_start
            while sub <= zone_end:
                txt = colA[sub]

                if txt == "BUSINESS TYPE BREAKDOWN":
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while e <= zone_end and isinstance(colA[e], str) and str(colA[e]).startswith("   "):
                        e += 1
                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)
                    sub = e
                    continue

                if isinstance(txt, str) and str(txt).startswith("TPTR Mode"):
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while e <= zone_end and not (
                        isinstance(colA[e], str)
                        and (
                            str(colA[e]).startswith("TPTR Mode")
                            or str(colA[e]) == "CN Status Breakdown"
                            or str(colA[e]) == "BUSINESS TYPE BREAKDOWN"
                            or str(colA[e]).startswith("NDR")
                        )
                    ):
                        e += 1
                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)
                    sub = e
                    continue

                if txt == "CN Status Breakdown":
                    ws.row_dimensions[sub].outline_level = 2
                    ws.row_dimensions[sub].collapsed = True

                    s = sub + 1
                    e = s
                    while e <= zone_end and isinstance(colA[e], str) and str(colA[e]).startswith("   "):
                        e += 1
                    if e - 1 >= s:
                        ws.row_dimensions.group(s, e - 1, hidden=True, outline_level=3)
                    sub = e
                    continue

                sub += 1

            row = r
        else:
            row += 1


def _format_summary_sheet(ws) -> None:
    """Same formatting as your original code."""
    if not _ensure_deps():
        st.stop()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="2F2F2F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 42
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r in range(2, ws.max_row + 1):
        c = ws[f"A{r}"]
        c.font = bold_font
        c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in range(2, ws.max_row + 1):
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center

    start_row = 2
    start_col = 2
    end_row = ws.max_row
    end_col = ws.max_column
    rng = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"

    rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(rng, rule)


# ============================================================
# ✅ PROCESSOR (OPTIMIZED)
# ============================================================
def process_edd_report(
    input_path: str,
    output_path: Optional[str] = None,
    source_sheet: str = "Query result",
    summary_sheet: str = "summary",
    apply_formatting: bool = True,
) -> str:
    if not _ensure_deps():
        st.stop()

    if output_path is None:
        file_path = input_path
    else:
        shutil.copyfile(input_path, output_path)
        file_path = output_path

    df = pd.read_excel(file_path, sheet_name=source_sheet)
    df.columns = df.columns.str.strip()

    date_cols = ["EDD_Date", "PICKUP_CHLN_DATE", "Reached At Destination", "DLY_Date"]
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    df = df.dropna(subset=["EDD_Date"])

    df["NEW_EDD_DATE"] = df["PICKUP_CHLN_DATE"] + pd.to_timedelta(df["TAT_DAYS"] - 1, unit="D")

    df["ON_TIME_ARRIVAL"] = "No"
    valid_arrival = (
        df["Reached At Destination"].notna()
        & df["EDD_Date"].notna()
        & df["NEW_EDD_DATE"].notna()
    )
    cond1 = df["Reached At Destination"] <= df["NEW_EDD_DATE"]
    cond2 = (
        (df["Reached At Destination"].dt.date == df["EDD_Date"].dt.date)
        & (df["Reached At Destination"].dt.time < time(12, 0))
    )
    df.loc[valid_arrival & (cond1 | cond2), "ON_TIME_ARRIVAL"] = "Yes"

    df["ON_TIME_DELIVERY"] = "No"
    df.loc[
        df["DLY_Date"].notna()
        & df["EDD_Date"].notna()
        & (df["DLY_Date"] <= df["EDD_Date"]),
        "ON_TIME_DELIVERY",
    ] = "Yes"

    df["Week_Label"] = "W-" + df["EDD_Date"].dt.isocalendar().week.astype(int).astype(str)

    summary = _build_summary_fast(df)

    with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        summary.to_excel(writer, sheet_name=summary_sheet)
        ws = writer.sheets[summary_sheet]
        _apply_row_grouping(ws)
        if apply_formatting:
            _format_summary_sheet(ws)

    return file_path


# ============================================================
# ✅ OpenAI helpers (lazy import = no startup crash)
# ============================================================
def get_openai_key() -> Optional[str]:
    key = None
    try:
        key = st.secrets.get("OPENAI_API_KEY", None)
    except Exception:
        key = None
    if not key:
        key = os.getenv("OPENAI_API_KEY") or os.getenv("OPENAI_SECRET_KEY")
    if not key:
        key = st.session_state.get("_openai_key")
    return key


def make_client() -> Tuple[Optional[object], Optional[str]]:
    key = get_openai_key()
    if not key:
        return None, "OpenAI key missing (add in Streamlit Secrets or sidebar)."

    try:
        from openai import OpenAI  # imported lazily
        return OpenAI(api_key=key), None
    except Exception as e:
        return None, f"OpenAI SDK import/init failed: {e}"


def llm_answer(client: object, model: str, system: str, user: str) -> str:
    last_err: Optional[Exception] = None

    # Responses API
    if hasattr(client, "responses"):
        try:
            resp = client.responses.create(
                model=model,
                instructions=system,
                input=user,
            )
            out = getattr(resp, "output_text", None)
            if out:
                return out.strip()
            return str(resp)
        except Exception as e:
            last_err = e

    # Chat Completions
    if hasattr(client, "chat") and hasattr(getattr(client, "chat"), "completions"):
        try:
            comp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user},
                ],
                temperature=0.2,
            )
            return (comp.choices[0].message.content or "").strip()
        except Exception as e:
            last_err = e

    return f"❌ OpenAI error: {last_err}" if last_err else "❌ OpenAI client not available."


# ============================================================
# ✅ Summary Q&A helpers (unchanged)
# ============================================================
@st.cache_data(show_spinner=False)
def load_summary_df(xlsx_bytes: bytes, sheet: str = "summary"):
    if not _ensure_deps():
        st.stop()
    bio = io.BytesIO(xlsx_bytes)
    df = pd.read_excel(bio, sheet_name=sheet, engine="openpyxl")
    if df.columns.size > 0 and str(df.columns[0]).lower().startswith("unnamed"):
        df = df.rename(columns={df.columns[0]: "Metric"})
    return df


def _normalize_week_label(week_num: int) -> str:
    return f"W-{int(week_num)}"


def extract_week(question: str) -> Optional[str]:
    q = (question or "").strip()

    m = re.search(r"\bW\s*[-_ ]\s*(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    m = re.search(r"\bweek\s*[-_ ]*\s*(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    m = re.search(r"\bweek(\d{1,2})\b", q, flags=re.IGNORECASE)
    if m:
        return _normalize_week_label(int(m.group(1)))

    return None


def zones_from_summary(summary_df) -> List[str]:
    if "Metric" not in summary_df.columns:
        return []
    zones = []
    for v in summary_df["Metric"].astype(str).tolist():
        m = re.match(r"^Picked Vol\. Zone (.*) %$", v.strip())
        if m:
            zones.append(m.group(1).strip())
    return sorted(set(zones))


def extract_zone(question: str, zones: List[str]) -> Optional[str]:
    q = (question or "").upper()
    zones_sorted = sorted(zones, key=lambda z: len(z), reverse=True)
    for z in zones_sorted:
        if z.upper() in q:
            return z
    return None


def select_relevant_rows(question: str, summary_df) -> List[str]:
    q = (question or "").lower()
    metric_col = "Metric" if "Metric" in summary_df.columns else summary_df.columns[0]
    metrics = summary_df[metric_col].astype(str).tolist()

    hits = []

    if "picked" in q and "volume" in q and "Picked Volume" in metrics:
        hits.append("Picked Volume")

    if "cn" in q and ("status" in q or "current" in q) and "CN Status Breakdown" in metrics:
        hits.append("CN Status Breakdown")

    if "ndr" in q and "NDR not available" in metrics:
        hits.append("NDR not available")

    if ("business" in q or "retail" in q or "scm" in q) and "BUSINESS TYPE BREAKDOWN" in metrics:
        hits.append("BUSINESS TYPE BREAKDOWN")

    if "tptr" in q or "mode" in q:
        for m in metrics:
            if str(m).startswith("TPTR Mode"):
                hits.append(m)
            if ("arrival" in q or "ota" in q) and "On Time Arrival" in str(m):
                hits.append(m)
            if ("delivery" in q or "otd" in q) and "On Time Delivery" in str(m):
                hits.append(m)

    if not hits:
        tokens = [t for t in re.findall(r"[a-zA-Z0-9]+", q) if len(t) > 2]
        stop = {"the","and","with","from","this","that","mein","me","ka","ki","ke","for","show","dikhao","bata","batao","please"}
        tokens = [t for t in tokens if t not in stop]
        for m in metrics:
            mm = str(m).lower()
            if any(t in mm for t in tokens):
                hits.append(m)

    seen = set()
    out = []
    for h in hits:
        if h not in seen and h in metrics:
            seen.add(h)
            out.append(h)
    return out


def answer_from_summary(summary_df, question: str) -> Tuple[Optional[str], Optional[str], List[Tuple[str, str]]]:
    if summary_df.empty:
        return None, None, []

    metric_col = "Metric" if "Metric" in summary_df.columns else summary_df.columns[0]

    zones = zones_from_summary(summary_df)
    week = extract_week(question)

    week_cols = [c for c in summary_df.columns if isinstance(c, str) and re.match(r"^W-\d{1,2}$", c.strip())]
    if not week and week_cols:
        week = sorted(week_cols, key=lambda x: int(x.split("-")[1]))[-1]

    zone = extract_zone(question, zones)
    if not zone:
        zone = "ALL INDIA" if "ALL INDIA" in zones else (zones[0] if zones else None)

    if not week or week not in summary_df.columns:
        return week, zone, []

    rows = select_relevant_rows(question, summary_df)

    result: List[Tuple[str, str]] = []
    for r in rows:
        match = summary_df[summary_df[metric_col] == r]
        if not match.empty:
            val = match.iloc[0][week]
            result.append((str(r), "" if pd.isna(val) else str(val)))

        if r in ("BUSINESS TYPE BREAKDOWN", "CN Status Breakdown"):
            start_idx = match.index[0] if not match.empty else None
            if start_idx is not None:
                for i in range(start_idx + 1, min(start_idx + 20, len(summary_df))):
                    mname = str(summary_df.iloc[i][metric_col])
                    if mname in ("BUSINESS TYPE BREAKDOWN", "CN Status Breakdown") or mname.startswith("TPTR Mode") or mname.startswith("Picked Vol. Zone"):
                        break
                    v = summary_df.iloc[i][week]
                    result.append((mname, "" if pd.isna(v) else str(v)))

    return week, zone, result


# ============================================================
# ✅ UI
# ============================================================
st.title("📦 EDD MIS Chatbot")
st.caption("Upload Excel → (1) View Summary  (2) Ask Questions  (3) Process Unprocessed File")

with st.sidebar:
    st.header("Settings")
    st.caption("If you still see 'Oh no', most likely requirements.txt is missing numpy/openpyxl.")
    model = st.selectbox("Model", options=["gpt-4o-mini", "gpt-4o"], index=0)

    st.markdown("### 🔑 OpenAI Secret Key")
    st.caption("Preferred: Streamlit Secrets → OPENAI_API_KEY")
    key_in = st.text_input("Paste key (sk-...)", type="password")
    if key_in:
        st.session_state["_openai_key"] = key_in

uploaded = st.file_uploader("Upload EDD Excel (.xlsx)", type=["xlsx"], key="main_upload")
tabs = st.tabs(["📊 Summary View", "💬 Ask a Question", "🛠️ Process File"])

with tabs[0]:
    if not uploaded:
        st.info("Upload a processed Excel first (jisme summary sheet already ho).")
    else:
        xbytes = uploaded.getvalue()
        try:
            sdf = load_summary_df(xbytes, sheet="summary")
            st.subheader("Summary Sheet")
            st.dataframe(sdf, use_container_width=True, height=600)
        except Exception as e:
            st.error(f"Summary sheet read nahi ho paayi: {e}")
            st.code(traceback.format_exc())

with tabs[1]:
    if not uploaded:
        st.info("Pehle Excel upload karo.")
    else:
        xbytes = uploaded.getvalue()
        try:
            summary_df = load_summary_df(xbytes, sheet="summary")
        except Exception as e:
            st.error(f"Summary sheet read nahi ho paayi: {e}")
            st.code(traceback.format_exc())
            summary_df = pd.DataFrame() if _ensure_deps() else None

        st.subheader("Ask Questions (Anything from Summary)")
        q = st.text_input(
            "Type your question:",
            placeholder="e.g. week-48 picked up volume / Week 51 ALL INDIA CN current status / week 46 ndr not available",
        )
        ask = st.button("Ask")

        if ask and q.strip() and summary_df is not None:
            week, zone, items = answer_from_summary(summary_df, q)

            if not items:
                st.error("Week/metric match nahi hua. Example: 'week-48 picked volume', 'Week 51 ALL INDIA CN current status'")
            else:
                client, cerr = make_client()
                context_lines = "\n".join([f"- {m}: {v}" for m, v in items])
                prompt = f"""Question: {q}

Excel Summary for {week} (zone={zone}):
{context_lines}

Give a short, direct answer. Do NOT show JSON or code. If multiple rows, format as bullet points."""
                if client:
                    with st.spinner("Generating answer..."):
                        ans = llm_answer(
                            client=client,
                            model=model,
                            system="You are a logistics MIS analyst. Answer only from the provided Excel summary context. If info missing, say what is missing.",
                            user=prompt,
                        )
                    if ans.strip().startswith("❌"):
                        st.error(ans)
                    else:
                        st.success(ans)
                else:
                    if cerr:
                        st.warning(cerr)
                    st.success("\n".join([f"• {m}: {v}" for m, v in items]))

with tabs[2]:
    st.subheader("Process Unprocessed File → Generate Summary")
    st.write("Yahan tum apna backend processing (pandas+openpyxl) run kara sakte ho.")
    unp = st.file_uploader("Upload Unprocessed Excel (.xlsx)", type=["xlsx"], key="unprocessed_upload")

    if unp:
        fast_mode = st.checkbox("⚡ Fast mode (skip formatting)", value=False)

        if st.button("Run Processing"):
            with st.spinner("Processing... (summary + grouping + formatting)"):
                with tempfile.TemporaryDirectory() as td:
                    in_path = os.path.join(td, "input.xlsx")
                    out_path = os.path.join(td, "EDD_Report_Processed.xlsx")
                    with open(in_path, "wb") as f:
                        f.write(unp.getvalue())

                    try:
                        process_edd_report(
                            input_path=in_path,
                            output_path=out_path,
                            source_sheet="Query result",
                            summary_sheet="summary",
                            apply_formatting=(not fast_mode),
                        )

                        with open(out_path, "rb") as f:
                            out_bytes = f.read()

                        st.success("✅ Processing done! Download processed file below.")
                        st.download_button(
                            "⬇️ Download Processed Excel",
                            data=out_bytes,
                            file_name="EDD_Report_Processed.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error(f"Processing failed: {e}")
                        st.code(traceback.format_exc())
